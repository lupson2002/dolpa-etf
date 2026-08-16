"""
core/emailer.py - whitelist Excel 생성 + Gmail SMTP 발송.
daily_batch 갱신 후 화이트리스트(KR/US)를 .xlsx(시트 2개) 첨부 이메일로 발송.

발신자 = EMAIL_USER(기본 yunjin.mike.choi@gmail.com), 수신자 = EMAIL_TO(동일).
Gmail 일반 비밀번호 대신 '앱 비밀번호' 16자리 사용(2차인증 필수).
"""
import logging
import math
import os
import re
import smtplib
from email.message import EmailMessage
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

# systemd ProtectSystem=strict 환경에서 openpyxl ExcelWriter 가 /tmp 임시파일을
# 못 만드는 현상 방지 — 쓰기 가능한 프로젝트 data/tmp 를 TMPDIR 로 강제 지정.
_DATA_TMP = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data", "tmp"
)
os.makedirs(_DATA_TMP, exist_ok=True)
os.environ.setdefault("TMPDIR", _DATA_TMP)

from config.settings import (
    EMAIL_HOST, EMAIL_PORT, EMAIL_USER, EMAIL_PASSWORD, EMAIL_TO,
)

logger = logging.getLogger(__name__)

# whitelist 항목 → Excel 컬럼 순서. (scanner.daily_batch 가 채운 필드)
COLUMNS = ["symbol", "name", "region", "score", "momentum", "bw40",
           "bb_upper", "donchian_high", "last_close"]

# 헤더 표시명(한글).
HEADERS = {
    "symbol": "종목코드", "name": "종목명", "region": "리전", "score": "최종점수",
    "momentum": "모멘텀", "bw40": "BW40", "bb_upper": "BB40상단",
    "donchian_high": "Donchian40고점", "last_close": "최근종가",
}


def build_whitelist_xlsx(whitelists: dict) -> bytes:
    """whitelists = {"KR": [...], "US": [...]} → xlsx 바이트(시트 KR/US 분리).

    각 시트는 score 내림차순 정렬. 빈 region도 빈 헤더 시트 생성.
    """
    wb = Workbook()
    first = True
    for region in ("KR", "US"):
        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = region
        # 헤더
        ws.append([HEADERS[c] for c in COLUMNS])
        # score 내림차순 정렬(원본 리스트 변경 방지 copy)
        rows = sorted(
            list(whitelists.get(region, [])),
            key=lambda x: x.get("score", 0.0) if x.get("score") is not None else 0.0,
            reverse=True,
        )
        for r in rows:
            ws.append([_cell(r.get(c, "")) for c in COLUMNS])
        # 컬럼 너비
        for i in range(1, len(COLUMNS) + 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 16
        # 헤더 볼드
        for i in range(1, len(COLUMNS) + 1):
            ws.cell(row=1, column=i).font = Font(bold=True)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cell(v):
    """값 정규화 — None/NaN → 빈 문자열, float 그대로."""
    if v is None:
        return ""
    try:
        if isinstance(v, float) and math.isnan(v):
            return ""
    except Exception:  # noqa: BLE001
        pass
    return v


def _strip_html(s: str) -> str:
    """텔레그램 HTML 포맷(<b> 등) → 순수 텍스트 (이메일 본문용)."""
    return re.sub(r"<[^>]+>", "", s or "")


def send_theme_email(theme_reports: dict) -> bool:
    """화이트리스트 테마 분석(단어/테마/연관성 + 워드클라우드) 이메일 발송.

    theme_reports = {"KR": {"text": str, "wordcloud": str|None},
                     "US": {"text": str, "wordcloud": str|None}}
    - text: 텔레그램 HTML 포맷의 테마/단어/연관성 섹션 (본문에서 태그 제거)
    - wordcloud: 워드클라우드 PNG 경로 (없으면 None — 이미지 첨부 생략)

    EMAIL_PASSWORD 미설정 시 스킵(경고 로그). 발송 실패 시 False — daily_batch
    흐름은 유지(Telegram 과 독립).
    """
    if not EMAIL_PASSWORD:
        logger.warning("EMAIL_PASSWORD 미설정 - 테마 분석 이메일 스킵 (.env 확인)")
        return False
    if not EMAIL_USER or not EMAIL_TO:
        logger.warning("EMAIL_USER/EMAIL_TO 미설정 - 테마 분석 이메일 스킵")
        return False

    body_lines = ["돌파etf 화이트리스트 테마 분석\n"]
    attachments: list[tuple[str, str]] = []  # (region, png_path)
    for region in ("US", "KR"):  # US 분석 먼저 → KR
        rep = theme_reports.get(region)
        if not rep or not rep.get("text"):
            continue
        body_lines.append(f"───── {region} ─────")
        body_lines.append(_strip_html(rep["text"]))
        body_lines.append("")
        wc = rep.get("wordcloud")
        if wc and os.path.exists(wc):
            attachments.append((region, wc))

    msg = EmailMessage()
    msg["Subject"] = "🧭 ETF 화이트리스트 테마 분석 (단어/연관성/워드클라우드)"
    msg["From"] = EMAIL_USER
    msg["To"] = EMAIL_TO
    msg.set_content("\n".join(body_lines))

    for region, wc_path in attachments:
        try:
            with open(wc_path, "rb") as f:
                img_bytes = f.read()
            msg.add_attachment(
                img_bytes,
                maintype="image",
                subtype="png",
                filename=f"wordcloud_{region}.png",
            )
        except Exception as e:  # noqa: BLE001
            logger.warning("워드클라우드 첨부 실패 %s: %s", wc_path, e)

    try:
        with smtplib.SMTP(EMAIL_HOST, EMAIL_PORT, timeout=30) as s:
            s.starttls()
            s.login(EMAIL_USER, EMAIL_PASSWORD)
            s.send_message(msg)
        logger.info("테마 분석 이메일 발송 성공 → %s (%d 첨부)",
                    EMAIL_TO, len(attachments))
        return True
    except Exception as e:  # noqa: BLE001
        logger.exception("테마 분석 이메일 발송 실패: %s", e)
        return False


def send_whitelist_email(whitelists: dict) -> bool:
    """whitelist KR/US 를 Excel 첨부해 EMAIL_TO 로 발송. 성공 시 True.

    EMAIL_PASSWORD(앱 비밀번호) 미설정 시 스킵(경고 로그). 발송 실패 시 False 반환 —
    daily_batch 흐름은 유지(Telegram 알림과 독립).
    """
    if not EMAIL_PASSWORD:
        logger.warning("EMAIL_PASSWORD 미설정 - 화이트리스트 이메일 스킵 (.env 확인)")
        return False
    if not EMAIL_USER or not EMAIL_TO:
        logger.warning("EMAIL_USER/EMAIL_TO 미설정 - 이메일 발송 스킵")
        return False

    kr_n = len(whitelists.get("KR", []))
    us_n = len(whitelists.get("US", []))

    try:
        xlsx = build_whitelist_xlsx(whitelists)
    except Exception as e:  # noqa: BLE001
        logger.exception("whitelist Excel 생성 실패: %s", e)
        return False

    msg = EmailMessage()
    msg["Subject"] = f"📊 ETF 화이트리스트 (KR {kr_n} / US {us_n})"
    msg["From"] = EMAIL_USER
    msg["To"] = EMAIL_TO
    msg.set_content(
        f"돌파etf 일일 화이트리스트\n"
        f"  KR: {kr_n}종목\n"
        f"  US: {us_n}종목\n\n"
        f"첨부 Excel(KR/US 시트 분리) 참고. score 내림차순 정렬."
    )
    msg.add_attachment(
        xlsx,
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="whitelist.xlsx",
    )

    try:
        with smtplib.SMTP(EMAIL_HOST, EMAIL_PORT, timeout=30) as s:
            s.starttls()
            s.login(EMAIL_USER, EMAIL_PASSWORD)
            s.send_message(msg)
        logger.info("화이트리스트 이메일 발송 성공 → %s (KR %d / US %d)",
                    EMAIL_TO, kr_n, us_n)
        return True
    except Exception as e:  # noqa: BLE001
        logger.exception("화이트리스트 이메일 발송 실패: %s", e)
        return False